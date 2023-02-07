from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title='Hcl'
#sheet["A1"].value=10
#sheet["B2"].value="test"
#sheet.cell(row=3,column=3).value="Hcl data"
j=0
for i in range(10,60,10):
    j+=1
    #sheet.cell(row=j,column=1).value=i
    sheet.cell(row=1,column=j).value=i
for cols in sheet.iter_cols(min_row=1,max_row=1,min_col=1,max_col=6):
    for r in cols:
        r.value=1
wb.save("C://hcl1/demo_open_xl_write.xlsx")