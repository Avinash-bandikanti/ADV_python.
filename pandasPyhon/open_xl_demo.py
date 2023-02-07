import openpyxl
wb=openpyxl.load_workbook("C://hcl1/emp.xlsx")
sheet=wb['Sheet1']
#print(sheet)
data=sheet['A1'].value
data1=sheet['A1:A10']
data2=sheet.cell(row=2,column=2).value
#print(data)
#print(data1)
#print(data2)
#Access all cells in row 1
#for rows in sheet.rows:
#    print([data.value for data in rows])
for i in range(2,5):
    print(sheet.cell(row=i,column=1).value)